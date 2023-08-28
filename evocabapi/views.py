import uuid
from datetime import datetime, date

from bson import json_util
from django.http import JsonResponse, Http404
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
#from django_rest.http import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import authentication, permissions
from django.contrib.auth.models import User
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from .models import WordsModel
from .serializers import WordsSerializer
from bson.objectid import ObjectId
import json
import openpyxl
from rest_framework.viewsets import ModelViewSet
from rest_framework.parsers import JSONParser


class ListUsers(APIView):
    """
    View to list all users in the system.

    * Requires token authentication.
    * Only admin users are able to access this view.
    """
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, format=None):
        """
        Return a list of all users.
        """
        usernames = [user.username for user in User.objects.all()]
        return Response(usernames)


class CustomAuthToken(ObtainAuthToken):

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data,
                                           context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email
        })

def importxls(request):
    if "GET" == request.method:
        return render(request, 'importxls.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        ii=0
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            print("row-------------------")
            excel_data.append(row_data)
            WordsModel.objects.create(word=row_data[0], translate=row_data[1],code=str(uuid.uuid4())[:16], trainDate='1970-01-01T00:00:00.000+00:00', train1=False,transcript=row_data[3],sound=row_data[5])
            #break

        return render(request, 'importxls.html', {"excel_data":excel_data})


class JSONEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, ObjectId):
            return str(o)
        return json.JSONEncoder.default(self, o)


class ViewWordsDel(APIView):
    model=WordsModel

    def get_object(self, pk):
        try:
            return WordsModel.objects.get(_id=ObjectId(str(pk)))
        except WordsModel.DoesNotExist:
            return Http404


    def get(self, request, *args, **kwargs):
        print("request=",request)
        print("nav=",request.GET.get("nav"))
        print("id=", request.GET.get("_id"))
        id = request.GET.get("_id")
        obj = self.get_object(id)
        querysetCount = WordsModel.objects.filter(trainDate='1970-01-01T00:00:00.000+00:00')
        #querysetCount = WordsModel.objects.filter(trainDate='1970-01-01T00:00:00.000+00:00').delete()
        #_id=64e4922ab48ef72e00eccac8
        if not querysetCount:
            count = 0
        else:
            count = querysetCount.count()

        for w in querysetCount:
            print(w.word)
        return Response({'word': count})

class ViewWords(APIView):
    model=WordsModel

    def get_object(self, pk):
        try:
            return WordsModel.objects.get(_id=ObjectId(str(pk)))
        except WordsModel.DoesNotExist:
            return Http404


    def get(self, request, *args, **kwargs):
        print("request=",request)
        print("nav=",request.GET.get("nav"))
        #print("nav=",request.data.get("nav"))
        print("id=",request.GET.get("_id"))
        navAction=""
        idAction=""
        nav = request.GET.get("nav")
        if nav is not None and nav != '':
            navAction=nav

        id=request.GET.get("_id")
        if id is not None and id != '':
            idAction=id


        #queryset = WordsModel.objects.order_by('trainDate').values()[:1] #filter(train1=True).values()
        if(navAction=="")or(idAction==""):
            queryset = WordsModel.objects.order_by('trainDate').values("word", "translate", "_id", "train1", "transcript", "sound", "trainDate")[:1] #filter(train1=True).values()
        elif (navAction=="prev"):
            #to=  WordsModel.objects.get(_id=id)
            obj = self.get_object(id)
            #to=WordsModel.objects.get(_id=ObjectId(str(id))).values("trainDate")
            print("quer=",obj)
            print("trainDate=",obj.trainDate)
            #queryset = WordsModel.objects.order_by('trainDate').values("word", "translate", "_id", "train1", "transcript", "sound", "trainDate").filter(trainDate__lt=obj.trainDate)[:1]  # filter(train1=True).values()
            queryset = WordsModel.objects.order_by('-trainDate').values("word", "translate", "_id", "train1", "transcript", "sound", "trainDate").filter(trainDate__lt=obj.trainDate)[:1]  # filter(train1=True).values()
            if not queryset:
                queryset = WordsModel.objects.order_by('-trainDate').values("word", "translate", "_id", "train1","transcript", "sound", "trainDate")[:1]
                print("queryset =",queryset)
        elif(navAction=="next"):
            obj = self.get_object(id)
            print("quer=", obj)
            print("trainDate=", obj.trainDate)
            # queryset = WordsModel.objects.order_by('trainDate').values("word", "translate", "_id", "train1", "transcript", "sound", "trainDate").filter(trainDate__lt=obj.trainDate)[:1]  # filter(train1=True).values()
            queryset = WordsModel.objects.order_by('trainDate').values("word", "translate", "_id", "train1",
                                                                        "transcript", "sound", "trainDate").filter(
                trainDate__gt=obj.trainDate)[:1]
            if not queryset:
                queryset = WordsModel.objects.order_by('trainDate').values("word", "translate", "_id", "train1",
                                                                            "transcript", "sound", "trainDate")[:1]
                print("queryset =", queryset)

        #!today_min = datetime.combine(datetime.date.today(), datetime.time.min)
        #!today_max = datetime.combine(datetime.date.today(), datetime.time.max)
        today_min = datetime.combine(timezone.now().date(), datetime.today().time().min)
        today_max = datetime.combine(timezone.now().date(), datetime.today().time().max)
        #today = date.today()
        #today = datetime.today()
        #querysetCount = WordsModel.objects.filter(trainDate__date=today)
        querysetCount = WordsModel.objects.filter(trainDate__range=(today_min, today_max), train1__in=[True])
        #querysetCount = WordsModel.objects.filter(train1__in=[False])
        if not querysetCount:
            count=0
        else:
            count = querysetCount.count()

        print("count=", count)

        #querysetCount=2;
        #querysetCountBad = WordsModel.objects.filter(trainDate__range=(today_min, today_max)).filter(train1__in=[False])
        querysetCountBad = WordsModel.objects.filter(trainDate__range=(today_min, today_max), train1__in=[False])

        if not querysetCountBad:
            countBad=0
        else:
            countBad=querysetCountBad.count()
            #countBad=0

        #querysetCountBad = WordsModel.objects.filter(train1=True).count()
        #queryset = queryset | queryset2;
        print("GET")
        print(queryset)
        trainDate = queryset[0]['trainDate']
        #df = df.iloc[:, 1:]
        queryset=json.loads(json_util.dumps(queryset))
        print("GET2")
        print(queryset)
        id=queryset[0]['_id']['$oid']

        #trainDate = list(WordsModel.objects.extra(select={'date': "to_char(<DATABASENAME>_<TableName>.created_at, 'YYYY-MM-DD hh:mi AM')"}).values_list('date', flat='true')
        print(id)
        #return Response({'word': queryset,'add':queryset2})
        #word=queryset[0];
        #print(word)
        return Response({'word': queryset[0]['word'], 'translate':queryset[0]['translate'],'id':id,'train1':queryset[0]['train1'], 'trainDate':trainDate,'transcript':queryset[0]['transcript'], 'sound':queryset[0]['sound'], 'countWord':count,'countWordBad':countBad})
        #return Response({'title': queryset})
        #queryset = WordsModel.objects.all()
        #words_serializer = WordsSerializer

    def post(self, request, format=None):
        #text=request.data.get("text")
        serializer = WordsSerializer(data=request.data)
        #obj = self.get_object(request.data.get("code"))
        #print(obj)
        #serializer = WordsSerializer(obj, data=request.data, partial=True)
        #print(serializer)
        #if serializer.is_valid(raise_exception=True):
        if serializer.is_valid():
            #serializer.update(obj, request.data)
            #obj.train1=True
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):
        print("PATCH")
        print(request.data)
        obj = self.get_object(request.data.get("id"))
        print("obj=")
        print(obj.word)

        serializer = WordsSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            #serializer.save()
            serializer.update(obj, request.data)
            #print(request.data.get("code"))
            #serializer.update(obj, request.data)
            #return Response(serializer.data, status=status.HTTP_202_ACCEPTED)
            return redirect('getword')
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        #train1=request.data.get("train1")
        #print(train1);
        #return Response("test", status.HTTP_201_CREATED)
        #serializer = WordsSerializer(data=request.data)
        #employee = get_object_or_404(Employee, id=employee_id)
        #if serializer.is_valid(raise_exception=True):
        #    serializer.update(serializer, request.data)
        #    return Response(serializer.data)
        #return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)